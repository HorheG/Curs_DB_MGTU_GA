import openpyxl
from openpyxl.styles import Alignment, Font
import os
import subprocess
import win32com.client as win32

def pdf_to(filename):
    path = os.getcwd()
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    workbook = excel.Workbooks.Open(path+r'\PDF\\'+filename)
    worksheet = workbook.Worksheets(1)
    pdf_path = path+r'\PDF\output.pdf'
    worksheet.ExportAsFixedFormat(0, pdf_path)
    workbook.Close()
    excel.Quit()
    subprocess.run([path+r'\PDF\output.pdf'], shell=True)
    
def save_billet(res):
    workbook = openpyxl.load_workbook('PDF/bilet.xlsx')
    worksheet = workbook['Sheet1']
    
    worksheet['A4'] = str(res[0])
    worksheet['B4'] = str(res[4])
    worksheet['E4'] = str(res[5])[:-3]
    worksheet['G4'] = str(res[7])
    worksheet['J4'] = str(res[8])[:-3]
    
    worksheet['L4'] = str(res[1])
    worksheet['M4'] = str(res[2])
    worksheet['N4'] = str(res[9])
    
    worksheet['F6'] = str(res[3])
    worksheet['F7'] = str(res[6])
    worksheet['B8'] = str(res[10]+' '+res[11]+' '+res[12])
    
    if res[14] == 'Паспорт':
        worksheet.merge_cells('B9:F9')
        worksheet['A9'] = str(res[14])
        worksheet['B9'] = str(res[15]) + ' ' + str(res[16])
    else:
        worksheet.merge_cells('A9:F9')
        worksheet.merge_cells('G9:J9')
        worksheet['A9'] = str(res[14])
        worksheet['G9'] = str(res[15]) + ' ' + str(res[16])

    workbook.save('PDF/output_bilet.xlsx')
    workbook.close()
    pdf_to('output_bilet.xlsx')
    
def train_route(res):
    workbook = openpyxl.load_workbook('PDF/route_train.xlsx')
    worksheet = workbook.active
    shift = 1
    for i in range(len(res)):

        if res[i][0] != res[i-1][0]:
            shift += 1
            worksheet.merge_cells('A'+str(i+ shift)+':E'+str(i + shift))
            cell = worksheet['A'+str(i+ shift)]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            worksheet['A'+str(i+shift)] = str(res[i][0])
            shift += 1
            list_cell = ['A'+str(i+shift), 'B'+str(i+shift),'C'+str(i+shift), 'D'+str(i+shift), 'E'+str(i+shift)]
            bold_style(list_cell, worksheet)
            center_style(list_cell, worksheet)
            
            worksheet['A'+str(i+shift)] = 'Станция'
            worksheet['B'+str(i+shift)] = 'Дата'
            worksheet['C'+str(i+shift)] = 'Время'
            worksheet['D'+str(i+shift)] = 'Стоянка'
            worksheet['E'+str(i+shift)] = 'Отправление'
            
            shift += 1
            list_cell = ['B'+str(i+shift),'C'+str(i+shift), 'D'+str(i+shift), 'E'+str(i+shift)]
            center_style(list_cell, worksheet)
            worksheet['A'+str(i+shift)] = res[i][1]
            worksheet['B'+str(i+shift)] = res[i][2]
            worksheet['C'+str(i+shift)] = ''
            worksheet['D'+str(i+shift)] = ''
            worksheet['E'+str(i+shift)] = res[i][5][:-3]
        else:
            list_cell = ['B'+str(i+shift),'C'+str(i+shift), 'D'+str(i+shift), 'E'+str(i+shift)]
            center_style(list_cell, worksheet)
            worksheet['A'+str(i+shift)] = res[i][1]
            worksheet['B'+str(i+shift)] = res[i][2]
            worksheet['C'+str(i+shift)] = res[i][3][:-3]
            worksheet['D'+str(i+shift)] = res[i][4][:-3]
            worksheet['E'+str(i+shift)] = res[i][5][:-3]
    
    workbook.save('PDF/train_route_out.xlsx')
    workbook.close()
    pdf_to('train_route_out.xlsx')

def bold_style(args, worksheet):
    for i in args:
        cell = worksheet[i]
        cell.font = Font(bold=True)
        
def center_style(args, worksheet):
    for i in args:
        cell = worksheet[i]
        cell.alignment = Alignment(horizontal='center', vertical='center')

'''
res = [('У128', 'Орск', '17.05.2023', '22:30:00', '00:00:00', '22:30:00'),
       ('У128', 'Медногорск', '17.05.2023', '23:07:00', '00:02:00', '23:09:00'), 
       ('У128', 'Саракташ', '18.05.2023', '00:59:00', '00:03:00', '01:02:00'), 
       ('У128', 'Оренбург', '18.05.2023', '02:29:00', '00:55:00', '03:24:00'), 
       ('У128', 'Новосерг', '18.05.2023', '05:16:00', '00:04:00', '05:20:00'),
       ('У128', 'Бузулук', '18.05.2023', '07:29:00', '00:23:00', '07:52:00'), 
       ('У128', 'Самара', '18.05.2023', '10:59:00', '00:46:00', '11:45:00'), 
       ('У128', 'Пенза', '18.05.2023', '18:32:00', '00:48:00', '19:20:00'), 
       ('У128', 'Ряжск 1', '19.05.2023', '03:10:00', '00:48:00', '03:58:00'), 
       ('У128', 'Москва каз', '19.05.2023', '09:52:00', '01:00:00', '10:52:00'), 
       ('Test1', 'Орск', '17.05.2023', '23:30:00', '00:00:00', '23:30:00'), 
       ('Test1', 'Медногорск', '17.05.2023', '00:07:00', '00:02:00', '00:09:00'),
       ('Test1', 'Саракташ', '18.05.2023', '01:59:00', '00:03:00', '02:02:00'), 
       ('Test1', 'Оренбург', '18.05.2023', '03:29:00', '00:55:00', '04:24:00'), 
       ('Test1', 'Новосерг', '18.05.2023', '06:16:00', '00:04:00', '06:20:00'), 
       ('Test1', 'Бузулук', '18.05.2023', '08:29:00', '00:23:00', '08:52:00'), 
       ('Test1', 'Самара', '18.05.2023', '11:59:00', '00:46:00', '12:45:00')]
'''
#train_route(res)
#res = ['Test1', 'СВ', 'Нижнее', 'Медногорск', '17-05-2023', '00:09:00', 'Самара', '17.05.2023', '11:59:00', '2250.00000', 'Гарипов', 'Фарит', 'Альбертович', '20.06.2000', 'Свидетельство о рождении', '5314', '568975']
#save_billet(res)
#pdf_to()
#Свидетельство о рождении